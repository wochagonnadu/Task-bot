"""Модуль для работы с Excel-отчетами"""

from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy import select, and_, func
from sqlalchemy.orm import selectinload, joinedload
from typing import Tuple, List, Dict

from src.database.db import get_session
from src.database.models import Task, Client, User, Project

class ExcelReportGenerator:
    """Генератор Excel-отчетов"""
    
    def __init__(self):
        self.wb = Workbook()
        # Удаляем лист по умолчанию, так как создадим свои
        self.wb.remove(self.wb.active)
        
    async def generate_report(self, period: str) -> bytes:
        """
        Генерация отчета за указанный период
        
        Args:
            period: Период (week/month/quarter/year)
            
        Returns:
            bytes: Excel-файл в виде байтов
        """
        date_from, date_to = self._get_period_dates(period)
        
        # Генерируем все вкладки
        await self._generate_clients_sheet(date_from, date_to)
        await self._generate_projects_sheet(date_from, date_to)
        await self._generate_employees_sheet(date_from, date_to)
        await self._generate_tasks_sheet(date_from, date_to)
        
        # Сохраняем в память
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _get_period_dates(self, period: str) -> Tuple[datetime, datetime]:
        """Получение дат начала и конца периода"""
        now = datetime.now()
        if period == 'week':
            date_from = now - timedelta(days=7)
        elif period == 'month':
            date_from = now - timedelta(days=30)
        elif period == 'quarter':
            date_from = now - timedelta(days=90)
        else:  # year
            date_from = now - timedelta(days=365)
        return date_from, now
    
    async def _generate_clients_sheet(self, date_from: datetime, date_to: datetime) -> None:
        """Создание вкладки с отчетом по клиентам"""
        sheet = self.wb.create_sheet("Клиенты")
        
        # Заголовки
        headers = [
            "Клиент",
            "Всего проектов",
            "Активные проекты",
            "Всего задач",
            "Выполнено задач",
            "Начало работы",
            "Последняя активность"
        ]
        
        self._write_headers(sheet, headers)
        
        # Получаем данные по клиентам
        async with get_session() as session:
            # Получаем клиентов с предзагруженными проектами и задачами
            stmt = (
                select(Client)
                .options(
                    selectinload(Client.projects),
                    selectinload(Client.tasks)
                )
            )
            result = await session.execute(stmt)
            clients = result.scalars().all()

            row = 2  # Начинаем после заголовков
            for client in clients:
                # Проекты уже загружены
                total_projects = len(client.projects)
                active_projects = len([p for p in client.projects if p.status == 'active'])
                
                # Фильтруем задачи по периоду
                tasks_list = [t for t in client.tasks if date_from <= t.created_at <= date_to]
                completed_tasks = len([t for t in tasks_list if t.status == Task.STATUS_COMPLETED])
                
                # Записываем данные
                data = [
                    client.name,
                    total_projects,
                    active_projects,
                    len(tasks_list),
                    completed_tasks,
                    client.created_at.strftime("%d.%m.%Y"),
                    client.updated_at.strftime("%d.%m.%Y")
                ]
                
                for col, value in enumerate(data, 1):
                    sheet.cell(row=row, column=col, value=value)
                row += 1
                
        self._adjust_column_width(sheet)
    
    async def _generate_projects_sheet(self, date_from: datetime, date_to: datetime) -> None:
        """Создание вкладки с отчетом по проектам"""
        sheet = self.wb.create_sheet("Проекты")
        
        headers = [
            "Проект",
            "Клиент",
            "Статус",
            "Дата начала",
            "Дата окончания",
            "Всего задач",
            "Выполнено задач",
            "Команда"
        ]
        
        self._write_headers(sheet, headers)
        
        # Получаем данные по проектам
        async with get_session() as session:
            # Загружаем проекты с клиентами и задачами
            stmt = (
                select(Project)
                .options(
                    joinedload(Project.client),
                    selectinload(Project.tasks).joinedload(Task.assignee)
                )
                .where(Project.created_at.between(date_from, date_to))
            )
            result = await session.execute(stmt)
            projects = result.unique().scalars().all()

            row = 2
            for project in projects:
                # Фильтруем задачи по периоду
                tasks_list = [t for t in project.tasks if date_from <= t.created_at <= date_to]
                completed_tasks = len([t for t in tasks_list if t.status == Task.STATUS_COMPLETED])
                
                # Собираем команду из исполнителей задач
                team = {t.assignee.full_name for t in tasks_list if t.assignee}
                
                data = [
                    project.name,
                    project.client.name if project.client else "Нет",
                    project.status,
                    project.start_date.strftime("%d.%m.%Y") if project.start_date else "Не указано",
                    project.end_date.strftime("%d.%m.%Y") if project.end_date else "Не указано",
                    len(tasks_list),
                    completed_tasks,
                    ", ".join(team) if team else "Нет исполнителей"
                ]
                
                for col, value in enumerate(data, 1):
                    sheet.cell(row=row, column=col, value=value)
                row += 1
                
        self._adjust_column_width(sheet)
    
    async def _generate_employees_sheet(self, date_from: datetime, date_to: datetime) -> None:
        """Создание вкладки с отчетом по сотрудникам"""
        sheet = self.wb.create_sheet("Сотрудники")
        
        headers = [
            "Сотрудник",
            "Всего задач",
            "Выполнено",
            "В работе",
            "Проекты"
        ]
        
        self._write_headers(sheet, headers)
        
        # Получаем данные по сотрудникам
        async with get_session() as session:
            # Получаем всех пользователей
            employees = await session.execute(
                select(User)
                .where(User.role == 'user')
            )
            employees = employees.scalars().all()

            row = 2
            for employee in employees:
                # Получаем задачи сотрудника за период
                tasks = await session.execute(
                    select(Task)
                    .options(joinedload(Task.project))
                    .where(
                        Task.assignee_id == employee.id,
                        Task.created_at.between(date_from, date_to)
                    )
                )
                tasks_list = tasks.unique().scalars().all()
                
                completed_tasks = len([t for t in tasks_list if t.status == Task.STATUS_COMPLETED])
                in_progress_tasks = len([t for t in tasks_list if t.status == Task.STATUS_IN_PROGRESS])
                
                # Собираем проекты
                projects = {t.project.name for t in tasks_list if t.project}
                
                data = [
                    employee.full_name or employee.username or f"User {employee.telegram_id}",
                    len(tasks_list),
                    completed_tasks,
                    in_progress_tasks,
                    ", ".join(projects) if projects else "Нет проектов"
                ]
                
                for col, value in enumerate(data, 1):
                    sheet.cell(row=row, column=col, value=value)
                row += 1
                
        self._adjust_column_width(sheet)

    async def _generate_tasks_sheet(self, date_from: datetime, date_to: datetime) -> None:
        """Создание вкладки со всеми задачами"""
        sheet = self.wb.create_sheet("Задачи")
        
        headers = [
            "Задача",
            "Описание",
            "Статус",
            "Исполнитель",
            "Клиент",
            "Проект",
            "Срок",
            "Создана",
            "Обновлена"
        ]
        
        self._write_headers(sheet, headers)
        
        async with get_session() as session:
            # Получаем все задачи с загрузкой связанных данных
            stmt = (
                select(Task)
                .options(
                    joinedload(Task.assignee),
                    joinedload(Task.client),
                    joinedload(Task.project)
                )
                .where(Task.created_at.between(date_from, date_to))
                .order_by(Task.created_at.desc())
            )
            tasks = await session.execute(stmt)
            tasks_list = tasks.unique().scalars().all()

            row = 2
            for task in tasks_list:
                data = [
                    task.title,
                    task.description or "",
                    task.status,
                    task.assignee.full_name if task.assignee else "Не назначен",
                    task.client.name if task.client else "Не указан",
                    task.project.name if task.project else "Не указан",
                    task.due_date.strftime("%d.%m.%Y") if task.due_date else "Не указан",
                    task.created_at.strftime("%d.%m.%Y"),
                    task.updated_at.strftime("%d.%m.%Y")
                ]
                
                for col, value in enumerate(data, 1):
                    sheet.cell(row=row, column=col, value=str(value))
                row += 1
                
        self._adjust_column_width(sheet)
    
    def _write_headers(self, sheet, headers: List[str]) -> None:
        """Запись и форматирование заголовков"""
        # Стили для заголовков
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color='CCE5FF',
            end_color='CCE5FF',
            fill_type='solid'
        )
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
    
    def _adjust_column_width(self, sheet) -> None:
        """Автоматическая подстройка ширины столбцов"""
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width